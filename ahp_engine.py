"""
AHP Engine — algoritma + Excel dengan formula asli (bukan angka mati)
"""
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

RI_TABLE = {1:0.00,2:0.00,3:0.58,4:0.90,5:1.12,
            6:1.24,7:1.32,8:1.41,9:1.45,10:1.49,
            11:1.51,12:1.54,13:1.56,14:1.57,15:1.59}

def get_ri(n): return RI_TABLE.get(n, 1.59)

# ── Core ──────────────────────────────────────────────────────────────────────

def compute_ahp_matrix(matrix: np.ndarray):
    n = matrix.shape[0]
    col_sums = matrix.sum(axis=0)
    norm     = matrix / col_sums
    priority = norm.mean(axis=1)
    Aw       = matrix @ priority
    lam      = Aw / priority
    lam_max  = lam.mean()
    CI  = (lam_max - n) / (n - 1) if n > 1 else 0.0
    RI  = get_ri(n)
    CR  = CI / RI if RI > 0 else 0.0
    return dict(n=n, matrix=matrix, col_sums=col_sums, norm=norm,
                priority=priority, lambda_vals=lam, lambda_max=lam_max,
                CI=CI, RI=RI, CR=CR, consistent=CR<0.10)

def compute_full_ahp(criteria_matrix, alt_matrices, crit_names, alt_names):
    crit_res   = compute_ahp_matrix(criteria_matrix)
    alt_results, alt_weight_cols = [], []
    for am in alt_matrices:
        r = compute_ahp_matrix(am)
        alt_results.append(r)
        alt_weight_cols.append(r["priority"])
    W      = np.column_stack(alt_weight_cols)
    scores = W @ crit_res["priority"]
    ranks  = scores.argsort()[::-1]
    rank_map = {int(idx): rank+1 for rank, idx in enumerate(ranks)}
    return dict(crit_res=crit_res, crit_weights=crit_res["priority"],
                alt_results=alt_results, W=W, scores=scores,
                rank_map=rank_map, crit_names=crit_names, alt_names=alt_names)

# ── Excel helpers — natural / human look ─────────────────────────────────────

def _b(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, txt, bold=True, bg=None, align="center"):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _num(ws, r, c, val_or_formula, fmt="0.0000", bold=False, bg=None):
    cell = ws.cell(row=r, column=c)
    cell.value          = val_or_formula
    cell.number_format  = fmt
    cell.font           = Font(name="Calibri", bold=bold, size=11)
    cell.alignment      = Alignment(horizontal="center", vertical="center")
    cell.border         = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _txt(ws, r, c, txt, bold=False, align="left"):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _b()
    return cell

def _sec(ws, r, c, txt, span):
    """Section title bar — grey, bold, left-aligned, no merge needed"""
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font  = Font(name="Calibri", bold=True, size=11)
    cell.fill  = PatternFill("solid", fgColor="D9D9D9")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _b()
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=c+span-1)

def _set_col_widths(ws, mapping):
    for col_letter, w in mapping.items():
        ws.column_dimensions[col_letter].width = w

# ── Build one AHP matrix sheet ────────────────────────────────────────────────

def _write_ahp_sheet(ws, res, labels, sheet_title):
    """
    Tulis satu sheet AHP lengkap dengan formula Excel.
    res    = dict dari compute_ahp_matrix
    labels = list nama baris/kolom (kriteria atau alternatif)
    """
    n = res["n"]
    mat = res["matrix"]

    # col widths
    ws.column_dimensions["A"].width = 18
    for i in range(2, n + 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

    R = 1  # current row pointer

    # ── Judul ──
    tc = ws.cell(row=R, column=1, value=sheet_title)
    tc.font = Font(name="Calibri", bold=True, size=12)
    tc.alignment = Alignment(horizontal="left")
    R += 2

    # ══ 1. Matriks Perbandingan Berpasangan ══════════════════════════════════
    _sec(ws, R, 1, "1. Matriks Perbandingan Berpasangan", n+2); R += 1

    # header row
    _hdr(ws, R, 1, "Kriteria", bg="EDEDED")
    for j, lbl in enumerate(labels):
        _hdr(ws, R, j+2, lbl, bg="EDEDED")
    R += 1

    # data rows — angka asli input (bukan formula, karena ini raw input)
    mat_start_row = R
    for i in range(n):
        _hdr(ws, R+i, 1, labels[i], bg="EDEDED")
        for j in range(n):
            v = mat[i, j]
            # tulis sebagai fraction-friendly number
            _num(ws, R+i, j+2, round(v, 8), fmt="0.######")

    # Total row — pakai formula SUM per kolom
    tot_row = mat_start_row + n
    _hdr(ws, tot_row, 1, "Total Kolom", bg="EDEDED")
    for j in range(n):
        col_l = get_column_letter(j+2)
        formula = f"=SUM({col_l}{mat_start_row}:{col_l}{mat_start_row+n-1})"
        _num(ws, tot_row, j+2, formula, fmt="0.0000", bold=True, bg="F2F2F2")

    R = tot_row + 2

    # ══ 2. Matriks Normalisasi ══════════════════════════════════════════════
    _sec(ws, R, 1, "2. Matriks Normalisasi (dibagi Total Kolom)", n+4); R += 1

    # header
    _hdr(ws, R, 1, "Kriteria", bg="EDEDED")
    for j, lbl in enumerate(labels):
        _hdr(ws, R, j+2, lbl, bg="EDEDED")
    _hdr(ws, R, n+2, "Jumlah Baris", bg="EDEDED")
    _hdr(ws, R, n+3, "Prioritas / Bobot", bg="EDEDED")
    _hdr(ws, R, n+4, "Eigen Value", bg="EDEDED")
    R += 1

    norm_start = R
    for i in range(n):
        _hdr(ws, R+i, 1, labels[i], bg="EDEDED")
        for j in range(n):
            # formula: mat_cell / total_cell
            mat_col  = get_column_letter(j+2)
            tot_col  = get_column_letter(j+2)
            mat_cell = f"{mat_col}{mat_start_row+i}"
            tot_cell = f"{tot_col}{tot_row}"
            _num(ws, R+i, j+2, f"={mat_cell}/{tot_cell}", fmt="0.000000")

        # Jumlah baris
        row_first = get_column_letter(2)
        row_last  = get_column_letter(n+1)
        _num(ws, R+i, n+2, f"=SUM({row_first}{R+i}:{row_last}{R+i})",
             fmt="0.000000")

        # Prioritas = jumlah baris / n
        jumlah_cell = f"{get_column_letter(n+2)}{R+i}"
        _num(ws, R+i, n+3, f"={jumlah_cell}/{n}", fmt="0.000000", bold=True, bg="EBF5EB")

    # Total prioritas
    prio_col  = get_column_letter(n+3)
    _hdr(ws, norm_start+n, 1, "Total", bg="EDEDED")
    _num(ws, norm_start+n, n+2,
         f"=SUM({get_column_letter(n+2)}{norm_start}:{get_column_letter(n+2)}{norm_start+n-1})",
         fmt="0.0000", bg="F2F2F2")
    _num(ws, norm_start+n, n+3,
         f"=SUM({prio_col}{norm_start}:{prio_col}{norm_start+n-1})",
         fmt="0.0000", bg="F2F2F2")

    # Eigen value = (A·w)/w — dihitung sebagai formula: (sum baris matriks × vektor prioritas) / prioritas_i
    for i in range(n):
        # Aw_i = Σ_j mat[i,j] * w_j
        terms = []
        for j in range(n):
            mat_c  = f"{get_column_letter(j+2)}{mat_start_row+i}"
            prio_c = f"{prio_col}{norm_start+j}"
            terms.append(f"{mat_c}*{prio_c}")
        aw_formula  = "=(" + "+".join(terms) + f")/{prio_col}{norm_start+i}"
        _num(ws, norm_start+i, n+4, aw_formula, fmt="0.000000", bg="EEF4FB")

    R = norm_start + n + 2

    # ══ 3. Uji Konsistensi ═══════════════════════════════════════════════════
    _sec(ws, R, 1, "3. Uji Konsistensi", n+4); R += 1

    # λ_max — nilai langsung (AVERAGE formula bisa error di locale Indonesia)
    _txt(ws, R, 1, "λ_max  =  rata-rata Eigen Value")
    lam_row = R
    _num(ws, R, n+2, round(res["lambda_max"], 8), fmt="0.000000"); R += 1

    # CI — formula aritmatika murni, aman semua locale
    _txt(ws, R, 1, f"CI  =  (λ_max - {n}) / ({n} - 1)")
    ci_row = R
    lam_cell = f"{get_column_letter(n+2)}{lam_row}"
    _num(ws, R, n+2, f"=({lam_cell}-{n})/({n}-1)", fmt="0.000000"); R += 1

    # RI — nilai literal tabel Saaty
    _txt(ws, R, 1, f"RI  (n={n}, Tabel Saaty)")
    ri_row = R
    _num(ws, R, n+2, get_ri(n), fmt="0.00"); R += 1

    # CR — formula aritmatika murni
    _txt(ws, R, 1, "CR  =  CI / RI")
    cr_row = R
    ci_cell = f"{get_column_letter(n+2)}{ci_row}"
    ri_cell = f"{get_column_letter(n+2)}{ri_row}"
    _num(ws, R, n+2, f"={ci_cell}/{ri_cell}", fmt="0.000000"); R += 1

    # Kesimpulan — nilai teks langsung (hindari IF lintas locale)
    _txt(ws, R, 1, "Kesimpulan")
    kl = ws.cell(row=R, column=n+2)
    kl.value = "KONSISTEN" if res["consistent"] else "TIDAK KONSISTEN"
    kl.font  = Font(name="Calibri", bold=True, size=11,
                    color="006100" if res["consistent"] else "9C0006")
    kl.fill  = PatternFill("solid", fgColor="EBF5EB" if res["consistent"] else "FFEBEE")
    kl.alignment = Alignment(horizontal="center", vertical="center")
    kl.border = _b()

    return R + 2

# ── Main export ───────────────────────────────────────────────────────────────

def generate_ahp_excel(full_res: dict) -> io.BytesIO:
    wb = Workbook()
    crit_names = full_res["crit_names"]
    alt_names  = full_res["alt_names"]
    n_crit     = len(crit_names)
    n_alt      = len(alt_names)

    # Sheet 1: Matriks Kriteria
    ws = wb.active
    ws.title = "Kriteria"
    _write_ahp_sheet(ws, full_res["crit_res"], crit_names,
                     "AHP – Matriks Perbandingan Kriteria")

    # Sheet per kriteria: matriks alternatif
    for j, (cn, ar) in enumerate(zip(crit_names, full_res["alt_results"])):
        ws2 = wb.create_sheet(title=f"Alt-{cn[:18]}")
        _write_ahp_sheet(ws2, ar, alt_names,
                         f"AHP – Alternatif terhadap Kriteria: {cn}")

    # Sheet Hasil Akhir
    ws_f = wb.create_sheet(title="Hasil Akhir")
    ws_f.column_dimensions["A"].width = 18
    for i in range(2, n_crit + 6):
        ws_f.column_dimensions[get_column_letter(i)].width = 14

    R = 1
    tc = ws_f.cell(row=R, column=1, value="AHP – Hasil Akhir & Peringkat")
    tc.font = Font(name="Calibri", bold=True, size=12)
    R += 2

    _sec(ws_f, R, 1, "Bobot Kriteria", n_crit+2); R += 1
    _hdr(ws_f, R, 1, "Kriteria", bg="EDEDED")
    _hdr(ws_f, R, 2, "Bobot", bg="EDEDED"); R += 1
    for ci, cn in enumerate(crit_names):
        _hdr(ws_f, R+ci, 1, cn, bg="EDEDED")
        ws_f.cell(row=R+ci, column=2).value = full_res["crit_weights"][ci]
        ws_f.cell(row=R+ci, column=2).number_format = "0.000000"
        ws_f.cell(row=R+ci, column=2).font   = Font(name="Calibri", size=11)
        ws_f.cell(row=R+ci, column=2).border = _b()
        ws_f.cell(row=R+ci, column=2).alignment = Alignment(horizontal="center")
    bobot_start = R
    _hdr(ws_f, R+n_crit, 1, "Total", bg="EDEDED")
    _num(ws_f, R+n_crit, 2,
         f"=SUM(B{bobot_start}:B{bobot_start+n_crit-1})", fmt="0.0000", bg="F2F2F2")
    R += n_crit + 2

    _sec(ws_f, R, 1, "Skor & Peringkat", n_crit+3); R += 1
    _hdr(ws_f, R, 1, "Alternatif", bg="EDEDED")
    for ci, cn in enumerate(crit_names):
        _hdr(ws_f, R, ci+2, cn, bg="EDEDED")
    _hdr(ws_f, R, n_crit+2, "Skor Total", bg="EDEDED")
    _hdr(ws_f, R, n_crit+3, "Peringkat",  bg="EDEDED"); R += 1

    # bobot per kriteria: simpan referensi sel
    skor_start = R
    W = full_res["W"]
    rank_map = full_res["rank_map"]
    ranking_order = sorted(range(n_alt), key=lambda i: rank_map[i])
    for ai in ranking_order:
        rank = rank_map[ai]
        bg_r = "EBF5EB" if rank == 1 else ("FFFDE7" if rank == 2 else None)
        _hdr(ws_f, R, 1, alt_names[ai], bg="EDEDED")
        for ci in range(n_crit):
            _num(ws_f, R, ci+2, round(W[ai, ci], 8), fmt="0.000000", bg=bg_r)
        # Skor: perkalian eksplisit tiap sel — menghindari #VALUE! dari SUMPRODUCT
        terms = [
            f"{get_column_letter(ci+2)}{R}*$B${bobot_start+ci}"
            for ci in range(n_crit)
        ]
        _num(ws_f, R, n_crit+2,
             "=" + "+".join(terms),
             fmt="0.000000", bold=True, bg=bg_r)
        # Peringkat — nilai literal agar tidak error di locale Indonesia
        rank_cell = ws_f.cell(row=R, column=n_crit+3)
        rank_cell.value     = rank
        rank_cell.font      = Font(name="Calibri", bold=True, size=11)
        rank_cell.alignment = Alignment(horizontal="center", vertical="center")
        rank_cell.border    = _b()
        if bg_r:
            rank_cell.fill = PatternFill("solid", fgColor=bg_r)
        R += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
