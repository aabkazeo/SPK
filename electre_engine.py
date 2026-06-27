"""
ELECTRE Engine — algoritma + Excel dengan formula asli, styling natural
"""
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# Core Algorithm
# ═══════════════════════════════════════════════════════════════════════════════

def normalize_matrix(X):
    col_norms = np.sqrt((X ** 2).sum(axis=0))
    col_norms[col_norms == 0] = 1
    return X / col_norms

def weighted_normalized(R, weights):
    return R * np.array(weights)

def concordance_sets(V, n_alt, criteria_types):
    sets = {}
    for k in range(n_alt):
        for l in range(n_alt):
            if k == l: continue
            c_set = []
            for j in range(len(criteria_types)):
                if criteria_types[j] == "benefit":
                    if V[k, j] >= V[l, j]: c_set.append(j)
                else:
                    if V[k, j] <= V[l, j]: c_set.append(j)
            sets[(k, l)] = c_set
    return sets

def discordance_sets(V, n_alt, criteria_types):
    c_sets = concordance_sets(V, n_alt, criteria_types)
    all_j = set(range(V.shape[1]))
    return {(k, l): sorted(all_j - set(c_sets[(k, l)]))
            for k in range(n_alt) for l in range(n_alt) if k != l}

def concordance_matrix(c_sets, weights, n_alt):
    C = np.full((n_alt, n_alt), np.nan)
    for k in range(n_alt):
        for l in range(n_alt):
            if k != l:
                C[k, l] = sum(weights[j] for j in c_sets[(k, l)])
    return C

def discordance_matrix(d_sets, V, n_alt, criteria_types):
    D = np.full((n_alt, n_alt), np.nan)
    n_crit = V.shape[1]
    for k in range(n_alt):
        for l in range(n_alt):
            if k == l: continue
            ds = d_sets[(k, l)]
            if not ds:
                D[k, l] = 0.0; continue
            num = max(abs(V[k, j] - V[l, j]) for j in ds)
            den = max(abs(V[k, j] - V[l, j]) for j in range(n_crit))
            D[k, l] = num / den if den != 0 else 0.0
    return D

def dominant_concordance(C, threshold_c, n_alt):
    F = np.full((n_alt, n_alt), np.nan)
    for k in range(n_alt):
        for l in range(n_alt):
            if k != l: F[k, l] = 1 if C[k, l] >= threshold_c else 0
    return F

def dominant_discordance(D, threshold_d, n_alt):
    G = np.full((n_alt, n_alt), np.nan)
    for k in range(n_alt):
        for l in range(n_alt):
            if k != l: G[k, l] = 1 if D[k, l] <= threshold_d else 0
    return G

def aggregate_dominance(F, G, n_alt):
    E = np.full((n_alt, n_alt), np.nan)
    for k in range(n_alt):
        for l in range(n_alt):
            if k != l: E[k, l] = 1 if (F[k, l] == 1 and G[k, l] == 1) else 0
    return E

def compute_ranking(C, D, n_alt):
    rows = []
    for k in range(n_alt):
        pairs = [(C[k, l], D[k, l]) for l in range(n_alt) if l != k]
        total = sum(c - d for c, d in pairs)
        rows.append({"alt_idx": k, "pairs": pairs, "E_score": total})
    scores = [r["E_score"] for r in rows]
    rank_map = {idx: rank + 1 for rank, idx in enumerate(np.argsort(scores)[::-1])}
    for r in rows: r["rank"] = rank_map[r["alt_idx"]]
    return rows

def run_electre(X, weights, criteria_types, alt_names, crit_names):
    n_alt, n_crit = X.shape
    R  = normalize_matrix(X)
    V  = weighted_normalized(R, weights)
    cs = concordance_sets(V, n_alt, criteria_types)
    ds = discordance_sets(V, n_alt, criteria_types)
    C  = concordance_matrix(cs, weights, n_alt)
    D  = discordance_matrix(ds, V, n_alt, criteria_types)
    off = [(k, l) for k in range(n_alt) for l in range(n_alt) if k != l]
    tc = np.mean([C[k, l] for k, l in off])
    td = np.mean([D[k, l] for k, l in off])
    F  = dominant_concordance(C, tc, n_alt)
    G  = dominant_discordance(D, td, n_alt)
    E  = aggregate_dominance(F, G, n_alt)
    rk = compute_ranking(C, D, n_alt)
    rank_map = {r["alt_idx"]: r["rank"] for r in rk}
    return dict(R=R, V=V, c_sets=cs, d_sets=ds, C=C, D=D,
                threshold_c=tc, threshold_d=td, F=F, G=G, E=E,
                ranking=rk, rank_map=rank_map,
                n_alt=n_alt, n_crit=n_crit,
                alt_names=alt_names, crit_names=crit_names,
                weights=weights, criteria_types=criteria_types, X=X)

# ═══════════════════════════════════════════════════════════════════════════════
# Excel Export — Natural styling, formula-based cells
# ═══════════════════════════════════════════════════════════════════════════════

def _b():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, txt, bold=True, bg=None):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _num(ws, r, c, val, fmt="0.0000", bold=False, bg=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.number_format = fmt
    cell.font          = Font(name="Calibri", bold=bold, size=11)
    cell.alignment     = Alignment(horizontal="center", vertical="center")
    cell.border        = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _txt(ws, r, c, txt, bold=False):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _b()
    return cell

def _sec(ws, r, c, txt, span):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=True, size=11)
    cell.fill      = PatternFill("solid", fgColor="D9D9D9")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _b()
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)

def _col(j): return get_column_letter(j)

def generate_electre_excel(res: dict) -> io.BytesIO:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Perhitungan ELECTRE"

    n_alt = res["n_alt"]; n_crit = res["n_crit"]
    alt   = res["alt_names"]; crit = res["crit_names"]
    X     = res["X"]; R_ = res["R"]; V_ = res["V"]
    cs    = res["c_sets"]; ds = res["d_sets"]
    C_    = res["C"]; D_ = res["D"]
    F_    = res["F"]; G_ = res["G"]; E_ = res["E"]

    # column widths
    ws.column_dimensions["A"].width = 20
    for i in range(2, n_crit + 12):
        ws.column_dimensions[_col(i)].width = 13

    ROW = 1  # row pointer

    # ── Title ──────────────────────────────────────────────────────────────────
    tc = ws.cell(row=ROW, column=1, value="Perhitungan Metode ELECTRE")
    tc.font = Font(name="Calibri", bold=True, size=13)
    ROW += 2

    # ══ 1. Matriks Keputusan Awal ══════════════════════════════════════════════
    _sec(ws, ROW, 1, "1. Matriks Keputusan Awal (X)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    x_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            _num(ws, ROW+i, j+2, X[i, j], fmt="0.##")
    # denominator row (sqrt of sum of squares) — pakai formula
    denom_row = x_start + n_alt
    _hdr(ws, denom_row, 1, "√(ΣX²)", bg="EDEDED")
    for j in range(n_crit):
        col_l = _col(j+2)
        formula = f"=SQRT(SUMSQ({col_l}{x_start}:{col_l}{x_start+n_alt-1}))"
        _num(ws, denom_row, j+2, formula, fmt="0.000000", bg="F2F2F2")
    ROW = denom_row + 2

    # ══ 2. Bobot & Jenis Kriteria ══════════════════════════════════════════════
    _sec(ws, ROW, 1, "2. Bobot dan Jenis Kriteria", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Kriteria", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    _hdr(ws, ROW, 1, "Bobot", bg="EDEDED")
    for j, w in enumerate(res["weights"]):
        _num(ws, ROW, j+2, w, fmt="0")
    w_row = ROW
    ROW += 1
    _hdr(ws, ROW, 1, "Jenis", bg="EDEDED")
    for j, t in enumerate(res["criteria_types"]):
        _txt(ws, ROW, j+2, t)
        ws.cell(row=ROW, column=j+2).alignment = Alignment(horizontal="center")
    ROW += 2

    # ══ 3. Matriks Normalisasi R ════════════════════════════════════════════════
    _sec(ws, ROW, 1, "3. Matriks Normalisasi (R)  —  r_ij = x_ij / √(Σx_ij²)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    r_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            x_cell    = f"{_col(j+2)}{x_start+i}"
            denom_cell = f"{_col(j+2)}{denom_row}"
            _num(ws, ROW+i, j+2, f"={x_cell}/{denom_cell}", fmt="0.00000000")
    ROW = r_start + n_alt + 1

    # ══ 4. Matriks Normalisasi Terbobot V ══════════════════════════════════════
    _sec(ws, ROW, 1, "4. Matriks Normalisasi Terbobot (V)  —  v_ij = w_j × r_ij", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    v_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            r_cell = f"{_col(j+2)}{r_start+i}"
            w_cell = f"{_col(j+2)}{w_row}"
            _num(ws, ROW+i, j+2, f"={r_cell}*{w_cell}", fmt="0.00000000")
    ROW = v_start + n_alt + 1

    # ══ 5. Himpunan Concordance & Discordance ══════════════════════════════════
    pairs = [(k, l) for k in range(n_alt) for l in range(n_alt) if k != l]
    _sec(ws, ROW, 1, "5. Himpunan Concordance dan Discordance", n_crit+3); ROW += 1
    _hdr(ws, ROW, 1, "Pasangan",     bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    _hdr(ws, ROW, n_crit+2, "C (Concordance)", bg="EDEDED")
    _hdr(ws, ROW, n_crit+3, "D (Discordance)", bg="EDEDED")
    ROW += 1
    for k, l in pairs:
        label = f"C{k+1}{l+1}"
        _hdr(ws, ROW, 1, label, bg="EDEDED")
        c_set = cs[(k, l)]; d_set = ds[(k, l)]
        for j in range(n_crit):
            val = 1 if j in c_set else 0
            bg  = "EBF5EB" if j in c_set else "FFEBEE"
            _num(ws, ROW, j+2, val, fmt="0", bg=bg)
        c_str = "{" + ",".join(str(jj+1) for jj in c_set) + "}" if c_set else "{ }"
        d_str = "{" + ",".join(str(jj+1) for jj in d_set) + "}" if d_set else "{ }"
        _txt(ws, ROW, n_crit+2, c_str)
        ws.cell(row=ROW, column=n_crit+2).alignment = Alignment(horizontal="center")
        _txt(ws, ROW, n_crit+3, d_str)
        ws.cell(row=ROW, column=n_crit+3).alignment = Alignment(horizontal="center")
        ROW += 1
    ROW += 1

    # ══ 6. Matriks Concordance C ═══════════════════════════════════════════════
    _sec(ws, ROW, 1, "6. Matriks Concordance (C)  —  c_kl = Σ w_j  (j ∈ Concordance)", n_alt+2); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for i, an in enumerate(alt): _hdr(ws, ROW, i+2, an, bg="EDEDED")
    _hdr(ws, ROW, n_alt+2, "Jumlah Baris", bg="EDEDED")
    ROW += 1
    c_mat_start = ROW
    for k, ak in enumerate(alt):
        _hdr(ws, ROW+k, 1, ak, bg="EDEDED")
        for l in range(n_alt):
            if k == l:
                _num(ws, ROW+k, l+2, "-", fmt="@", bg="F2F2F2")
            else:
                _num(ws, ROW+k, l+2, C_[k, l], fmt="0.0000")
        # jumlah baris (skip diagonal "-")
        parts = [f"{_col(l+2)}{ROW+k}" for l in range(n_alt) if l != k]
        _num(ws, ROW+k, n_alt+2, "=" + "+".join(parts), fmt="0.0000", bg="F2F2F2")
    # threshold = AVERAGE of all off-diagonal
    # collect all off-diagonal cells
    all_off = [f"{_col(l+2)}{c_mat_start+k}"
               for k in range(n_alt) for l in range(n_alt) if k != l]
    threshold_c_row = ROW + n_alt
    # write threshold label + formula
    _txt(ws, threshold_c_row, 1, "Threshold  c̄  =  rata-rata nilai C", bold=True)
    tc_cell_addr = f"{_col(n_alt+2)}{threshold_c_row}"
    _num(ws, threshold_c_row, n_alt+2,
         "=(" + "+".join(all_off) + f")/{len(all_off)}", fmt="0.0000", bold=True, bg="FFF9C4")
    ROW = threshold_c_row + 2

    # ══ 7. Matriks Discordance D ═══════════════════════════════════════════════
    _sec(ws, ROW, 1, "7. Matriks Discordance (D)", n_alt+2); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for i, an in enumerate(alt): _hdr(ws, ROW, i+2, an, bg="EDEDED")
    _hdr(ws, ROW, n_alt+2, "Jumlah Baris", bg="EDEDED")
    ROW += 1
    d_mat_start = ROW
    for k, ak in enumerate(alt):
        _hdr(ws, ROW+k, 1, ak, bg="EDEDED")
        for l in range(n_alt):
            if k == l:
                _num(ws, ROW+k, l+2, "-", fmt="@", bg="F2F2F2")
            else:
                _num(ws, ROW+k, l+2, D_[k, l], fmt="0.00000000")
        parts_d = [f"{_col(l+2)}{ROW+k}" for l in range(n_alt) if l != k]
        _num(ws, ROW+k, n_alt+2, "=" + "+".join(parts_d), fmt="0.00000000", bg="F2F2F2")
    all_off_d = [f"{_col(l+2)}{d_mat_start+k}"
                 for k in range(n_alt) for l in range(n_alt) if k != l]
    threshold_d_row = ROW + n_alt
    _txt(ws, threshold_d_row, 1, "Threshold  d̄  =  rata-rata nilai D", bold=True)
    td_cell_addr = f"{_col(n_alt+2)}{threshold_d_row}"
    _num(ws, threshold_d_row, n_alt+2,
         "=(" + "+".join(all_off_d) + f")/{len(all_off_d)}", fmt="0.00000000", bold=True, bg="FFF9C4")
    ROW = threshold_d_row + 2

    # ══ 8. Matriks Dominan F (Concordance) ════════════════════════════════════
    _sec(ws, ROW, 1,
         f"8. Matriks Dominan Concordance (F)  —  1 jika C_kl ≥ c̄, else 0",
         n_alt+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for i, an in enumerate(alt): _hdr(ws, ROW, i+2, an, bg="EDEDED")
    ROW += 1
    f_mat_start = ROW
    for k in range(n_alt):
        _hdr(ws, ROW+k, 1, alt[k], bg="EDEDED")
        for l in range(n_alt):
            if k == l:
                _num(ws, ROW+k, l+2, "-", fmt="@", bg="F2F2F2")
            else:
                c_cell = f"{_col(l+2)}{c_mat_start+k}"
                formula = f"=IF({c_cell}>={tc_cell_addr},1,0)"
                bg = "EBF5EB" if F_[k, l] == 1 else None
                _num(ws, ROW+k, l+2, formula, fmt="0", bg=bg)
    ROW = f_mat_start + n_alt + 1

    # ══ 9. Matriks Dominan G (Discordance) ════════════════════════════════════
    _sec(ws, ROW, 1,
         f"9. Matriks Dominan Discordance (G)  —  1 jika D_kl ≤ d̄, else 0",
         n_alt+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for i, an in enumerate(alt): _hdr(ws, ROW, i+2, an, bg="EDEDED")
    ROW += 1
    g_mat_start = ROW
    for k in range(n_alt):
        _hdr(ws, ROW+k, 1, alt[k], bg="EDEDED")
        for l in range(n_alt):
            if k == l:
                _num(ws, ROW+k, l+2, "-", fmt="@", bg="F2F2F2")
            else:
                d_cell = f"{_col(l+2)}{d_mat_start+k}"
                formula = f"=IF({d_cell}<={td_cell_addr},1,0)"
                bg = "EBF5EB" if G_[k, l] == 1 else None
                _num(ws, ROW+k, l+2, formula, fmt="0", bg=bg)
    ROW = g_mat_start + n_alt + 1

    # ══ 10. Matriks Agregate E ════════════════════════════════════════════════
    _sec(ws, ROW, 1,
         "10. Matriks Agregate Dominan (E)  —  1 jika F_kl=1 AND G_kl=1",
         n_alt+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for i, an in enumerate(alt): _hdr(ws, ROW, i+2, an, bg="EDEDED")
    ROW += 1
    e_mat_start = ROW
    for k in range(n_alt):
        _hdr(ws, ROW+k, 1, alt[k], bg="EDEDED")
        for l in range(n_alt):
            if k == l:
                _num(ws, ROW+k, l+2, "-", fmt="@", bg="F2F2F2")
            else:
                f_cell = f"{_col(l+2)}{f_mat_start+k}"
                g_cell = f"{_col(l+2)}{g_mat_start+k}"
                formula = f"=IF(AND({f_cell}=1,{g_cell}=1),1,0)"
                bg = "EBF5EB" if E_[k, l] == 1 else None
                _num(ws, ROW+k, l+2, formula, fmt="0", bg=bg)
    ROW = e_mat_start + n_alt + 1

    # ══ 11. Hasil Akhir ════════════════════════════════════════════════════════
    _sec(ws, ROW, 1, "11. Skor Akhir dan Peringkat", 6); ROW += 1
    for ci, h in enumerate(["Alternatif", "Σ C_kl", "Σ D_kl", "Skor E = ΣC − ΣD", "Peringkat"], 1):
        _hdr(ws, ROW, ci, h, bg="EDEDED")
    ROW += 1
    skor_start = ROW
    ranking_sorted = sorted(res["ranking"], key=lambda x: x["rank"])
    for rd in ranking_sorted:
        ai   = rd["alt_idx"]
        rank = rd["rank"]
        bg_r = "EBF5EB" if rank == 1 else ("FFFDE7" if rank == 2 else None)
        _hdr(ws, ROW, 1, alt[ai], bg="EDEDED")
        # ΣC_kl = sum of row k in C matrix (excluding diagonal)
        c_parts = [f"{_col(l+2)}{c_mat_start+ai}" for l in range(n_alt) if l != ai]
        d_parts = [f"{_col(l+2)}{d_mat_start+ai}" for l in range(n_alt) if l != ai]
        sum_c_formula = "=" + "+".join(c_parts)
        sum_d_formula = "=" + "+".join(d_parts)
        sc_cell = f"B{ROW}"; sd_cell = f"C{ROW}"
        _num(ws, ROW, 2, sum_c_formula, fmt="0.0000", bg=bg_r)
        _num(ws, ROW, 3, sum_d_formula, fmt="0.00000000", bg=bg_r)
        _num(ws, ROW, 4, f"={sc_cell}-{sd_cell}", fmt="0.0000", bold=True, bg=bg_r)
        # Peringkat — nilai literal agar tidak error di locale Indonesia
        rk_cell = ws.cell(row=ROW, column=5)
        rk_cell.value     = rank
        rk_cell.font      = Font(name="Calibri", bold=True, size=11)
        rk_cell.alignment = Alignment(horizontal="center", vertical="center")
        rk_cell.border    = _b()
        if bg_r: rk_cell.fill = PatternFill("solid", fgColor=bg_r if bg_r else "FFFFFF")
        ROW += 1

    ws.freeze_panes = "B2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
