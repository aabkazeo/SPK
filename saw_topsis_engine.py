"""
SAW & TOPSIS Engine — formula-based Excel, natural styling
"""
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _col(j): return get_column_letter(j)

def _b():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, txt, bold=True, bg=None):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _num(ws, r, c, val, fmt="0.0000", bold=False, bg=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.number_format = fmt
    cell.font          = Font(name="Calibri", bold=bold, size=11)
    cell.alignment     = Alignment(horizontal="center", vertical="center")
    cell.border        = _b()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _txt(ws, r, c, txt, bold=False, align="left"):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=bold, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _b()
    return cell

def _sec(ws, r, c, txt, span):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(name="Calibri", bold=True, size=11)
    cell.fill      = PatternFill("solid", fgColor="D9D9D9")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _b()
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)

def _set_widths(ws, n_crit):
    ws.column_dimensions["A"].width = 18
    for i in range(2, n_crit + 8):
        ws.column_dimensions[_col(i)].width = 13

# ═══════════════════════════════════════════════════════════════════════════════
# SAW
# ═══════════════════════════════════════════════════════════════════════════════

def run_saw(X, weights, criteria_types, alt_names, crit_names):
    n_alt, n_crit = X.shape
    R = np.zeros_like(X, dtype=float)
    for j in range(n_crit):
        if criteria_types[j] == "benefit":
            mx = X[:, j].max()
            R[:, j] = X[:, j] / mx if mx else 0
        else:
            mn = X[:, j].min()
            R[:, j] = mn / X[:, j] if X[:, j].all() else 0
    w      = np.array(weights, dtype=float)
    w_norm = w / w.sum()
    scores = R @ w_norm
    order  = scores.argsort()[::-1]
    rank_map = {int(idx): rank+1 for rank, idx in enumerate(order)}
    return dict(X=X, R=R, w_norm=w_norm, scores=scores, rank_map=rank_map,
                alt_names=alt_names, crit_names=crit_names,
                weights=weights, criteria_types=criteria_types)

def generate_saw_excel(res) -> io.BytesIO:
    wb = Workbook(); ws = wb.active; ws.title = "Perhitungan SAW"
    alt    = res["alt_names"]; crit = res["crit_names"]
    n_alt  = len(alt); n_crit = len(crit)
    X      = res["X"]; w_norm = res["w_norm"]
    criteria_types = res["criteria_types"]
    _set_widths(ws, n_crit)

    ROW = 1
    tc = ws.cell(row=ROW, column=1, value="Perhitungan Metode SAW (Simple Additive Weighting)")
    tc.font = Font(name="Calibri", bold=True, size=13)
    ROW += 2

    # ── 1. Matriks keputusan awal ──────────────────────────────────────────────
    _sec(ws, ROW, 1, "1. Matriks Keputusan Awal (X)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    x_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            _num(ws, ROW+i, j+2, X[i, j], fmt="0.##")
    ROW = x_start + n_alt + 1

    # ── 2. Bobot ──────────────────────────────────────────────────────────────
    _sec(ws, ROW, 1, "2. Bobot dan Jenis Kriteria", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Bobot Asli", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    _hdr(ws, ROW, 1, "Nilai Bobot", bg="EDEDED")
    for j, w in enumerate(res["weights"]):
        _num(ws, ROW, j+2, w, fmt="0")
    w_raw_row = ROW; ROW += 1
    _hdr(ws, ROW, 1, "Jenis Kriteria", bg="EDEDED")
    for j, t in enumerate(criteria_types):
        _txt(ws, ROW, j+2, t)
        ws.cell(row=ROW, column=j+2).alignment = Alignment(horizontal="center")
    ROW += 1
    # bobot ternormalisasi — pakai formula
    _hdr(ws, ROW, 1, "Bobot Ternormalisasi", bg="EDEDED")
    total_w_parts = "+".join([f"{_col(j+2)}{w_raw_row}" for j in range(n_crit)])
    w_norm_row = ROW
    for j in range(n_crit):
        _num(ws, ROW, j+2,
             f"={_col(j+2)}{w_raw_row}/({total_w_parts})",
             fmt="0.000000")
    ROW += 2

    # ── 3. Nilai max/min per kolom ─────────────────────────────────────────────
    _sec(ws, ROW, 1, "3. Nilai MAX dan MIN per Kriteria (untuk normalisasi)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    _hdr(ws, ROW, 1, "MAX", bg="EDEDED")
    max_row = ROW
    for j in range(n_crit):
        col_l = _col(j+2)
        _num(ws, ROW, j+2,
             f"=MAX({col_l}{x_start}:{col_l}{x_start+n_alt-1})",
             fmt="0.####", bg="EBF5EB")
    ROW += 1
    _hdr(ws, ROW, 1, "MIN", bg="EDEDED")
    min_row = ROW
    for j in range(n_crit):
        col_l = _col(j+2)
        _num(ws, ROW, j+2,
             f"=MIN({col_l}{x_start}:{col_l}{x_start+n_alt-1})",
             fmt="0.####", bg="FFEBEE")
    ROW += 2

    # ── 4. Matriks Normalisasi R ───────────────────────────────────────────────
    _sec(ws, ROW, 1,
         "4. Matriks Normalisasi (R)  —  Benefit: x/MAX, Cost: MIN/x",
         n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    r_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            x_cell   = f"{_col(j+2)}{x_start+i}"
            max_cell = f"{_col(j+2)}{max_row}"
            min_cell = f"{_col(j+2)}{min_row}"
            if criteria_types[j] == "benefit":
                formula = f"={x_cell}/{max_cell}"
            else:
                formula = f"={min_cell}/{x_cell}"
            _num(ws, ROW+i, j+2, formula, fmt="0.000000")
    ROW = r_start + n_alt + 1

    # ── 5. Nilai Preferensi V ──────────────────────────────────────────────────
    _sec(ws, ROW, 1,
         "5. Nilai Preferensi (V)  —  V_i = Σ w_j × r_ij",
         n_crit+2); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    _hdr(ws, ROW, n_crit+2, "V (Skor Akhir)", bg="EDEDED")
    ROW += 1
    v_start = ROW
    rank_map = res["rank_map"]
    order = sorted(range(n_alt), key=lambda i: rank_map[i])
    for i, ai in enumerate(order):
        rank = rank_map[ai]
        bg_r = "EBF5EB" if rank == 1 else ("FFFDE7" if rank == 2 else None)
        _hdr(ws, ROW+i, 1, alt[ai], bg="EDEDED")
        for j in range(n_crit):
            r_cell = f"{_col(j+2)}{r_start+ai}"
            w_cell = f"{_col(j+2)}{w_norm_row}"
            _num(ws, ROW+i, j+2, f"={r_cell}*{w_cell}", fmt="0.000000", bg=bg_r)
        # V_i = perkalian eksplisit (SUMPRODUCT bisa #VALUE! di locale Indonesia)
        terms_v = [
            f"{_col(j+2)}{ROW+i}*{_col(j+2)}{w_norm_row}"
            for j in range(n_crit)
        ]
        _num(ws, ROW+i, n_crit+2,
             "=" + "+".join(terms_v),
             fmt="0.000000", bold=True, bg=bg_r)
    ROW = v_start + n_alt + 1

    # ── 6. Ranking ────────────────────────────────────────────────────────────
    _sec(ws, ROW, 1, "6. Peringkat Akhir", n_crit+3); ROW += 1
    for ci, h in enumerate(["Alternatif", "Skor V", "Peringkat"], 1):
        _hdr(ws, ROW, ci, h, bg="EDEDED")
    ROW += 1
    skor_start = ROW
    for i, ai in enumerate(order):
        rank = rank_map[ai]
        bg_r = "EBF5EB" if rank == 1 else ("FFFDE7" if rank == 2 else None)
        _hdr(ws, ROW+i, 1, alt[ai], bg="EDEDED")
        # referensikan ke kolom V di atas
        v_cell = f"{_col(n_crit+2)}{v_start+i}"
        _num(ws, ROW+i, 2, f"={v_cell}", fmt="0.000000", bg=bg_r)
        # Peringkat — nilai literal agar tidak error di locale Indonesia
        rk_ai = rank_map[ai]
        rk_cell = ws.cell(row=ROW+i, column=3)
        rk_cell.value     = rk_ai
        rk_cell.font      = Font(name="Calibri", bold=True, size=11)
        rk_cell.alignment = Alignment(horizontal="center", vertical="center")
        rk_cell.border    = _b()
        if bg_r: rk_cell.fill = PatternFill("solid", fgColor=bg_r)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════════════════════════════
# TOPSIS
# ═══════════════════════════════════════════════════════════════════════════════

def run_topsis(X, weights, criteria_types, alt_names, crit_names):
    n_alt, n_crit = X.shape
    w      = np.array(weights, dtype=float)
    w_norm = w / w.sum()
    norms  = np.sqrt((X**2).sum(axis=0)); norms[norms==0] = 1
    R      = X / norms
    V      = R * w_norm
    A_pos  = np.array([V[:, j].max() if criteria_types[j]=="benefit" else V[:, j].min()
                       for j in range(n_crit)])
    A_neg  = np.array([V[:, j].min() if criteria_types[j]=="benefit" else V[:, j].max()
                       for j in range(n_crit)])
    D_pos  = np.sqrt(((V - A_pos)**2).sum(axis=1))
    D_neg  = np.sqrt(((V - A_neg)**2).sum(axis=1))
    denom  = D_pos + D_neg; denom[denom==0] = 1e-10
    C      = D_neg / denom
    order  = C.argsort()[::-1]
    rank_map = {int(idx): rank+1 for rank, idx in enumerate(order)}
    return dict(X=X, R=R, V=V, w_norm=w_norm,
                A_pos=A_pos, A_neg=A_neg,
                D_pos=D_pos, D_neg=D_neg, C=C,
                rank_map=rank_map,
                alt_names=alt_names, crit_names=crit_names,
                weights=weights, criteria_types=criteria_types)

def generate_topsis_excel(res) -> io.BytesIO:
    wb = Workbook(); ws = wb.active; ws.title = "Perhitungan TOPSIS"
    alt    = res["alt_names"]; crit = res["crit_names"]
    n_alt  = len(alt); n_crit = len(crit)
    X      = res["X"]; criteria_types = res["criteria_types"]
    _set_widths(ws, n_crit)

    ROW = 1
    tc = ws.cell(row=ROW, column=1, value="Perhitungan Metode TOPSIS")
    tc.font = Font(name="Calibri", bold=True, size=13)
    ROW += 2

    # ── 1. Matriks awal ────────────────────────────────────────────────────────
    _sec(ws, ROW, 1, "1. Matriks Keputusan Awal (X)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Alternatif", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    x_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            _num(ws, ROW+i, j+2, X[i, j], fmt="0.##")
    ROW = x_start + n_alt + 1

    # ── 2. Bobot ──────────────────────────────────────────────────────────────
    _sec(ws, ROW, 1, "2. Bobot dan Jenis Kriteria", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "Bobot", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    _hdr(ws, ROW, 1, "Bobot Asli", bg="EDEDED")
    for j, w in enumerate(res["weights"]): _num(ws, ROW, j+2, w, fmt="0")
    w_raw_row = ROW; ROW += 1
    _hdr(ws, ROW, 1, "Jenis Kriteria", bg="EDEDED")
    for j, t in enumerate(criteria_types):
        _txt(ws, ROW, j+2, t); ws.cell(row=ROW, column=j+2).alignment = Alignment(horizontal="center")
    ROW += 1
    _hdr(ws, ROW, 1, "Bobot Ternormalisasi", bg="EDEDED")
    total_w_parts = "+".join([f"{_col(j+2)}{w_raw_row}" for j in range(n_crit)])
    w_norm_row = ROW
    for j in range(n_crit):
        _num(ws, ROW, j+2,
             f"={_col(j+2)}{w_raw_row}/({total_w_parts})",
             fmt="0.000000")
    ROW += 2

    # ── 3. Normalisasi Vektor R ────────────────────────────────────────────────
    _sec(ws, ROW, 1,
         "3. Matriks Normalisasi Vektor (R)  —  r_ij = x_ij / √(Σ x_ij²)",
         n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    # denom: sqrt(SUMSQ per kolom)
    _hdr(ws, ROW, 1, "√(Σ x_ij²)", bg="EDEDED")
    denom_row = ROW
    for j in range(n_crit):
        col_l = _col(j+2)
        _num(ws, ROW, j+2,
             f"=SQRT(SUMSQ({col_l}{x_start}:{col_l}{x_start+n_alt-1}))",
             fmt="0.000000", bg="F2F2F2")
    ROW += 1
    r_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            x_cell = f"{_col(j+2)}{x_start+i}"
            d_cell = f"{_col(j+2)}{denom_row}"
            _num(ws, ROW+i, j+2, f"={x_cell}/{d_cell}", fmt="0.00000000")
    ROW = r_start + n_alt + 1

    # ── 4. Normalisasi Terbobot V ──────────────────────────────────────────────
    _sec(ws, ROW, 1,
         "4. Matriks Normalisasi Terbobot (V)  —  v_ij = w_j × r_ij",
         n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    v_start = ROW
    for i, an in enumerate(alt):
        _hdr(ws, ROW+i, 1, an, bg="EDEDED")
        for j in range(n_crit):
            r_cell = f"{_col(j+2)}{r_start+i}"
            w_cell = f"{_col(j+2)}{w_norm_row}"
            _num(ws, ROW+i, j+2, f"={r_cell}*{w_cell}", fmt="0.00000000")
    ROW = v_start + n_alt + 1

    # ── 5. Solusi Ideal A+ dan A– ─────────────────────────────────────────────
    _sec(ws, ROW, 1, "5. Solusi Ideal Positif (A+) dan Negatif (A–)", n_crit+1); ROW += 1
    _hdr(ws, ROW, 1, "", bg="EDEDED")
    for j, cn in enumerate(crit): _hdr(ws, ROW, j+2, cn, bg="EDEDED")
    ROW += 1
    _hdr(ws, ROW, 1, "A+ (Ideal Positif)", bg="EDEDED")
    apos_row = ROW
    for j in range(n_crit):
        col_l = _col(j+2)
        if criteria_types[j] == "benefit":
            formula = f"=MAX({col_l}{v_start}:{col_l}{v_start+n_alt-1})"
        else:
            formula = f"=MIN({col_l}{v_start}:{col_l}{v_start+n_alt-1})"
        _num(ws, ROW, j+2, formula, fmt="0.00000000", bg="EBF5EB")
    ROW += 1
    _hdr(ws, ROW, 1, "A– (Ideal Negatif)", bg="EDEDED")
    aneg_row = ROW
    for j in range(n_crit):
        col_l = _col(j+2)
        if criteria_types[j] == "benefit":
            formula = f"=MIN({col_l}{v_start}:{col_l}{v_start+n_alt-1})"
        else:
            formula = f"=MAX({col_l}{v_start}:{col_l}{v_start+n_alt-1})"
        _num(ws, ROW, j+2, formula, fmt="0.00000000", bg="FFEBEE")
    ROW += 2

    # ── 6. Jarak D+ dan D– ────────────────────────────────────────────────────
    _sec(ws, ROW, 1,
         "6. Jarak ke Solusi Ideal  —  D+ = √Σ(v_ij−A+)², D– = √Σ(v_ij−A–)²",
         n_crit+3); ROW += 1
    for ci, h in enumerate(["Alternatif"] +
                             [f"(v−A+)² {cn}" for cn in crit] +
                             ["D+", "D–", "Ci", "Peringkat"], 1):
        _hdr(ws, ROW, ci, h, bg="EDEDED")
    ROW += 1
    dplus_col  = n_crit + 2
    dminus_col = n_crit + 3
    ci_col     = n_crit + 4
    rank_col   = n_crit + 5
    d_start    = ROW
    rank_map   = res["rank_map"]
    order      = sorted(range(n_alt), key=lambda i: rank_map[i])
    for idx, ai in enumerate(order):
        rank = rank_map[ai]
        bg_r = "EBF5EB" if rank == 1 else ("FFFDE7" if rank == 2 else None)
        _hdr(ws, ROW+idx, 1, alt[ai], bg="EDEDED")
        # (v_ij - A+)^2 per kolom — needed for SQRT(SUM(...))
        sq_pos_refs = []
        sq_neg_refs = []
        for j in range(n_crit):
            v_cell    = f"{_col(j+2)}{v_start+ai}"
            apos_cell = f"{_col(j+2)}{apos_row}"
            aneg_cell = f"{_col(j+2)}{aneg_row}"
            sq_pos    = f"({v_cell}-{apos_cell})^2"
            sq_neg    = f"({v_cell}-{aneg_cell})^2"
            # tulis nilai (v-A+)^2 sebagai formula di kolom j+2
            _num(ws, ROW+idx, j+2, f"={sq_pos}", fmt="0.00000000", bg=bg_r)
            sq_pos_refs.append(f"{_col(j+2)}{ROW+idx}")
            sq_neg_refs.append(sq_neg)  # untuk D– pakai formula langsung

        # D+
        _num(ws, ROW+idx, dplus_col,
             f"=SQRT({'+'.join(sq_pos_refs)})",
             fmt="0.00000000", bold=True, bg=bg_r)
        # D–
        _num(ws, ROW+idx, dminus_col,
             f"=SQRT({'+'.join(sq_neg_refs)})",
             fmt="0.00000000", bold=True, bg=bg_r)
        # Ci = D– / (D+ + D–)
        dp_cell = f"{_col(dplus_col)}{ROW+idx}"
        dm_cell = f"{_col(dminus_col)}{ROW+idx}"
        _num(ws, ROW+idx, ci_col,
             f"={dm_cell}/({dp_cell}+{dm_cell})",
             fmt="0.000000", bold=True, bg=bg_r)
        # Peringkat — nilai literal agar tidak error di locale Indonesia
        rk_cell = ws.cell(row=ROW+idx, column=rank_col)
        rk_cell.value     = rank
        rk_cell.font      = Font(name="Calibri", bold=True, size=11)
        rk_cell.alignment = Alignment(horizontal="center", vertical="center")
        rk_cell.border    = _b()
        if bg_r: rk_cell.fill = PatternFill("solid", fgColor=bg_r)
    ROW = d_start + n_alt

    ws.freeze_panes = "B2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
